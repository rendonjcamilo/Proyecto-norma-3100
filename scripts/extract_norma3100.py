#!/usr/bin/env python3
"""
Extract Norma 3100 data from Excel and generate JSON model.

This script reads the Archivo_Consolidaddo_Resolucion_3100-2019.xlsx file and
creates a JSON model with:
1. 7 transversal standards with all their criteria and sub-criteria
2. 38 health services with their specific criteria
3. Complete normative text for each criterion

Usage:
    python scripts/extract_norma3100.py \
      --input "docs/Norma 3100/Archivo_Consolidaddo_Resolucion_3100-2019.xlsx" \
      --output docs/norma3100-model.json
"""

import json
import argparse
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import re


class Norma3100Extractor:
    """Extract Norma 3100 data from Excel workbook."""

    # Mapping of sheet names to standard codes
    TRANSVERSAL_STANDARDS = {
        "11.1.1TH": ("TSTH", "Talento Humano"),
        "11.1.2.INF": ("TSINF", "Información"),
        "11.1.3.DOT": ("TSDOT", "Dotación"),
        "11.1.4MD": ("TSMD", "Medicamentos y Dispositivos"),
        "11.1.5.PP": ("TSPP", "Procesos Prioritarios"),
        "11.1.6.HCR": ("TSHCR", "Historia Clínica y Registros"),
        "11.1.7INT": ("TSINT", "Integralidad"),
    }

    # Service groups and their services
    SERVICE_GROUPS = {
        "11.2": {
            "name": "Consulta Externa",
            "services": {
                "11.2.1.S_CE_G": ("CEG", "Consulta Externa General"),
                "11.2.2.S_CE_E": ("CEE", "Consulta Externa Especializada"),
                "11.2.3.S_CE_V": ("CEV", "Consulta Externa Virtual"),
                "11.2.4.S_CE_SST": ("CES", "Consulta Externa Seguridad Social en el Trabajo"),
            },
        },
        "11.3": {
            "name": "Apoyo Diagnóstico y Complementación Terapéutica",
            "services": {
                "11.3.1.S_TR": ("TRF", "Terapia Física"),
                "11.3.2.SF": ("TER", "Terapias Especializadas"),
                "11.3.3.S_Rx_OD": ("RXO", "Rayos X Odontológicos"),
                "11.3.4.S_IDx": ("IDX", "Imágenes Diagnósticas"),
                "11.3.5.S_MNuc": ("MNX", "Medicina Nuclear"),
                "11.3.6.S_Radio": ("RDT", "Radioterapia"),
                "11.3.7.S_Quimio": ("QMT", "Quimioterapia"),
                "11.3.8 S_Dx VASC": ("DVX", "Diagnóstico Vascular"),
                "11.3.9S_HI": ("HTR", "Hematología"),
                "11.3.10 S_GPT": ("GNT", "Genética"),
                "11.3.11.S_TMLC": ("TLC", "Técnicas Mínimamente Invasivas Laparoscópicas"),
                "11.3.12.S_LC": ("LAB", "Laboratorio Clínico"),
                "11.3.13 S_TM_CU": ("TMC", "Terapias Manuales y Corporales"),
                "11.3.14.S_LCU": ("LAC", "Laboratorio Clínico Urgencias"),
                "11.3.15.S_LHT": ("LHT", "Laboratorio de Hematología"),
                "11.3.16.S_LPT": ("LPT", "Laboratorio de Patología"),
                "11.3.17.S_Dial": ("DLS", "Diálisis"),
            },
        },
        "11.4": {
            "name": "Internación",
            "services": {
                "11.4.1.S_HP": ("HGP", "Hospitalización General"),
                "11.4.2.S_HP_PC": ("HPP", "Hospitalización Pediatría"),
                "11.4.3.S_CBN": ("OBN", "Obstetricia Neonatal"),
                "11.4.4.S_CIN": ("CIN", "Cuidado Intensivo Neonatal"),
                "11.4.5.S_CINN": ("CII", "Cuidado Intermedio Neonatal"),
                "11.4.6. S_CINP": ("CIP", "Cuidado Intensivo Pediátrico"),
                "11.4.7.S_CIP": ("CIM", "Cuidado Intermedio Pediátrico"),
                "11.4.8.S_CIMA": ("CIA", "Cuidado Intensivo Adulto"),
                "11.4.9.S_CIA": ("CIM", "Cuidado Intermedio Adulto"),
                "11.4.10.S_HSM CSP": ("HSC", "Hospitalización Salud Mental"),
                "11.4.11.S_HSP": ("HSP", "Hospitalización Psiquiátrica"),
                "11.4.12.S_CB_CSP": ("CPC", "Cuidado Básico Psiquiátrico"),
            },
        },
        "11.5": {
            "name": "Quirúrgico",
            "services": {
                "11.5.1.S_CX": ("QRG", "Quirúrgico"),
            },
        },
        "11.6": {
            "name": "Atención Inmediata",
            "services": {
                "11.6.1.S_UR": ("URG", "Urgencias"),
                "11.6.2.S_TR_AS": ("TAS", "Transporte Asistencial"),
                "11.6.3.S_AT_PH": ("APH", "Atención Prehospitalaria"),
                "11.6.4.S_A.parto": ("APR", "Atención del Parto"),
            },
        },
    }

    def __init__(self, xlsx_path: Path):
        """Initialize the extractor with an Excel file path."""
        self.xlsx_path = xlsx_path
        self.wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        self.model: Dict[str, Any] = {
            "version": "resolucion-3100-2019",
            "source": "Archivo_Consolidaddo_Resolucion_3100-2019.xlsx",
            "standards": [],
            "service_groups": [],
        }

    def extract_criteria_from_sheet(
        self, sheet: Worksheet, standard_code: str = None
    ) -> List[Dict[str, Any]]:
        """
        Extract criteria from a sheet.

        Args:
            sheet: Worksheet to extract from
            standard_code: Standard code (for transversal standards)

        Returns:
            List of criteria dictionaries
        """
        criteria = []
        current_section = None
        current_subsection = None

        # Find header row (usually row 3 or 4 with "Estandar", "Criterios", "Estado", "Comentarios")
        header_row = None
        for row_num in range(1, 10):
            row = list(sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
            if row and isinstance(row[0], str) and "standar" in row[0].lower():
                header_row = row_num
                break

        if not header_row:
            return []

        # Extract criteria starting after header
        for row_num, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
        ):
            if not row or not row[0]:  # Skip empty rows
                continue

            codigo = row[0]
            criterio_text = row[1] if len(row) > 1 else None
            estado = row[2] if len(row) > 2 else None
            comentarios = row[3] if len(row) > 3 else None

            if not criterio_text:
                continue

            criterio_text = str(criterio_text).strip()

            # Section headers (Estado is None/empty)
            if not estado:
                # This is a section header for grouping
                if re.match(r'^Complejidad|^Modalidad|^Categoria', criterio_text):
                    current_section = criterio_text
                    continue
                # Check if it's a numbering (starts with digit and dot)
                if re.match(r'^\d+\.', criterio_text):
                    current_subsection = criterio_text.split('.')[0]
                else:
                    current_section = criterio_text
                continue

            # Extract criterion number from text
            number_match = re.match(r'^(\d+(?:\.\d+)*)', criterio_text)
            number = number_match.group(1) if number_match else None

            # Determine complexity
            complexity = "simple"
            if "media" in criterio_text.lower() or "mediana" in criterio_text.lower():
                complexity = "medium"
            elif "alta" in criterio_text.lower() or "alta_complejidad" in criterio_text.lower():
                complexity = "high"

            # Generate criterion code
            if standard_code:
                criterion_code = f"{standard_code}-{len(criteria) + 1:03d}"
            else:
                criterion_code = None

            criterion = {
                "code": criterion_code,
                "number": number,
                "text": criterio_text,
                "complexity": complexity,
                "is_mandatory": estado != "NA",
                "state": estado,  # C, NC, or NA
            }

            if current_section:
                criterion["section"] = current_section

            criteria.append(criterion)

        return criteria

    def extract_transversal_standards(self) -> List[Dict[str, Any]]:
        """Extract all 7 transversal standards with their criteria."""
        standards = []

        for sheet_name, (code, name) in self.TRANSVERSAL_STANDARDS.items():
            if sheet_name not in self.wb.sheetnames:
                print(f"Warning: Sheet '{sheet_name}' not found")
                continue

            sheet = self.wb[sheet_name]
            criteria = self.extract_criteria_from_sheet(sheet, code)

            # Renumber criteria codes sequentially
            for i, criterion in enumerate(criteria, 1):
                criterion["code"] = f"{code}-{i:03d}"

            standard = {
                "code": code,
                "name": name,
                "is_transversal": True,
                "sheet_ref": sheet_name,
                "criteria_count": len(criteria),
                "criteria": criteria,
            }

            standards.append(standard)
            print(f"[OK] {code}: {name} ({len(criteria)} criteria)")

        return standards

    def extract_services(self) -> List[Dict[str, Any]]:
        """Extract all services with their specific criteria."""
        service_groups = []

        for group_code, group_info in self.SERVICE_GROUPS.items():
            group = {
                "group_code": group_code,
                "group_name": group_info["name"],
                "services": [],
            }

            for sheet_name, (service_code, service_name) in group_info["services"].items():
                if sheet_name not in self.wb.sheetnames:
                    print(f"Warning: Sheet '{sheet_name}' not found")
                    continue

                sheet = self.wb[sheet_name]

                # Extract service-specific criteria
                service_standards = self._extract_service_standards(sheet, service_code)

                service = {
                    "code": service_code,
                    "name": service_name,
                    "sheet_ref": sheet_name,
                    "applicable_transversal_standards": [
                        "TSTH", "TSINF", "TSDOT", "TSMD", "TSPP", "TSHCR", "TSINT"
                    ],
                    "specific_standards": service_standards,
                }

                # Check if multi-instance (like Diagnóstico Vascular)
                if sheet_name == "11.3.8 S_Dx VASC":
                    service["multi_instance"] = True
                    service["max_instances"] = 7
                    service["instance_label"] = "Sede de diagnóstico vascular"

                group["services"].append(service)
                total_criteria = sum(
                    len(ss["criteria"]) for ss in service_standards
                )
                print(f"  [OK] {service_code}: {service_name} ({total_criteria} criteria)")

            if group["services"]:
                service_groups.append(group)

        return service_groups

    def _extract_service_standards(
        self, sheet: Worksheet, service_code: str
    ) -> List[Dict[str, Any]]:
        """Extract service-specific standards (by subject area) from a sheet."""
        service_standards_dict: Dict[str, Dict[str, Any]] = {}

        # Find header row
        header_row = None
        for row_num in range(1, 10):
            row = list(sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
            if row and isinstance(row[0], str) and "standar" in row[0].lower():
                header_row = row_num
                break

        if not header_row:
            return []

        # Extract criteria
        for row_num, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
        ):
            if not row or not row[0]:
                continue

            service_std_code = str(row[0]).strip()
            criterio_text = row[1] if len(row) > 1 else None
            estado = row[2] if len(row) > 2 else None

            if not criterio_text or not service_std_code:
                continue

            criterio_text = str(criterio_text).strip()

            # Check if this is a criterion (starts with a number) or a section header
            has_number = re.match(r'^(\d+(?:\.\d+)*)', criterio_text)
            is_section = re.match(r'^(Complejidad|Modalidad|Categoria)', criterio_text)

            # Skip pure section headers (no number)
            if not has_number and is_section:
                continue

            # Initialize service standard if not exists
            if service_std_code not in service_standards_dict:
                # Determine parent transversal standard
                parent = self._map_to_transversal(service_std_code)
                service_standards_dict[service_std_code] = {
                    "code": service_std_code,
                    "parent_transversal": parent,
                    "criteria": [],
                }

            # Extract criterion number
            number_match = re.match(r'^(\d+(?:\.\d+)*)', criterio_text)
            number = number_match.group(1) if number_match else None

            criterion = {
                "number": number,
                "text": criterio_text,
                "is_mandatory": estado != "NA" if estado else True,
                "state": estado,
            }

            service_standards_dict[service_std_code]["criteria"].append(criterion)

        return list(service_standards_dict.values())

    def _map_to_transversal(self, service_std_code: str) -> Optional[str]:
        """Map a service-specific standard code to its parent transversal standard."""
        mapping = {
            # Talento Humano
            "_TH": "TSTH",
            # Infraestructura / Información
            "_INF": "TSINF",
            # Dotación
            "_DOT": "TSDOT",
            # Medicamentos y Dispositivos
            "_MD": "TSMD",
            # Procesos Prioritarios
            "_PP": "TSPP",
            # Historia Clínica y Registros
            "_HCR": "TSHCR",
            # Integralidad
            "_INT": "TSINT",
        }

        for suffix, parent in mapping.items():
            if service_std_code.endswith(suffix):
                return parent

        # Default mapping by position in service code
        if "_" in service_std_code:
            suffix = service_std_code.split("_")[-1]
            # Map based on letter patterns
            if suffix.startswith("TH"):
                return "TSTH"
            elif suffix.startswith("INF"):
                return "TSINF"
            elif suffix.startswith("DOT"):
                return "TSDOT"
            elif suffix.startswith("MD"):
                return "TSMD"
            elif suffix.startswith("PP"):
                return "TSPP"
            elif suffix.startswith("HCR"):
                return "TSHCR"
            elif suffix.startswith("INT"):
                return "TSINT"

        return "TSTH"  # Default fallback

    def extract(self) -> Dict[str, Any]:
        """Run the complete extraction process."""
        print("\n" + "=" * 70)
        print("EXTRACTING NORMA 3100 FROM EXCEL")
        print("=" * 70 + "\n")

        print("Extracting 7 Transversal Standards:")
        print("-" * 70)
        self.model["standards"] = self.extract_transversal_standards()

        print("\nExtracting Service-Specific Criteria:")
        print("-" * 70)
        self.model["service_groups"] = self.extract_services()

        # Calculate and add statistics
        self.model["statistics"] = {
            "total_transversal_standards": len(self.model["standards"]),
            "total_transversal_criteria": sum(
                s["criteria_count"] for s in self.model["standards"]
            ),
            "total_service_groups": len(self.model["service_groups"]),
            "total_services": sum(
                len(sg["services"]) for sg in self.model["service_groups"]
            ),
        }

        return self.model


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Extract Norma 3100 data from Excel and generate JSON model"
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Path to the Excel file",
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Path to output JSON file",
    )

    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        print(f"Error: Input file not found: {input_path}")
        return 1

    # Create output directory if needed
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Extract data
    extractor = Norma3100Extractor(input_path)
    model = extractor.extract()

    # Write JSON
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(model, f, indent=2, ensure_ascii=False)

    print("\n" + "=" * 70)
    print("EXTRACTION COMPLETE")
    print("=" * 70)
    print(f"\nJSON model saved to: {output_path}")
    print(f"\nStatistics:")
    for key, value in model["statistics"].items():
        print(f"  {key}: {value}")

    return 0


if __name__ == "__main__":
    exit(main())
